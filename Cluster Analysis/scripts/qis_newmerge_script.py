# -*- coding: utf-8 -*-
"""
Created on Fri Jun 30 15:36:04 2023

@author: HQP6CBS
"""

import pandas as pd
import os
from functools import reduce
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
#from EquityHedging.datamanager import data_manager as dm
CWD = os.getcwd()


# Replace 'strategies.xlsx' with the path to your original Excel file
existing_file = 'QIS Universe Time Series.xlsx'
existing_sheet_name = 'CS'

# Replace 'new_data.xlsx' with the path to your new Excel file
new_file = 'update_QIS_uni.xlsx'
new_sheet_name = 'Sheet1'

#a = dm.get_qis_uni_dict()
# Read the original Excel file into a Pandas DataFrame
existing_df = pd.read_excel(existing_file, sheet_name=existing_sheet_name, skiprows=1, engine='openpyxl')

# Read the new data Excel file into a Pandas DataFrame
new_df = pd.read_excel(new_file, sheet_name=new_sheet_name)


# Get the ticker names from the second row of the new DataFrame (B2, C2, D2, etc.)
ticker_names_new = new_df.columns[1::2].tolist()

ticker_names_new = [f"{value} Index" for value in ticker_names_new]

# Rename the columns in the new DataFrame by adding the ' Index' suffix
new_df.rename(columns=dict(zip(new_df.columns[1::2], ticker_names_new)), inplace=True)

# Get the ticker names from the second row of the existing DataFrame (B2, C2, D2, etc.)
ticker_names_existing = existing_df.columns[2:].tolist()

# Count the number of tickers in the new data that match existing tickers
common_tickers_count = len(set(ticker_names_new).intersection(ticker_names_existing))

print("Number of tickers in the new data that match existing tickers:", common_tickers_count)

#Create a dictionary to store DataFrames for each ticker in the new data
new_strategy_data_frames = {}

# Iterate through the new ticker names and create separate DataFrames for each ticker
for ticker in ticker_names_new:
    # Extract date and price columns for the current ticker in the new data
    date_column = new_df.iloc[2:, 0]
    price_column = new_df[ticker][2:]
    
    # Create a new DataFrame for the current ticker in the new data
    ticker_df = pd.DataFrame({'Date': date_column, ticker: price_column})
    
    # Convert the 'Date' column to datetime type
    ticker_df['Date'] = pd.to_datetime(ticker_df['Date'])
    
    # Add the DataFrame to the dictionary with ticker as the key
    new_strategy_data_frames[ticker] = ticker_df

# Merge all DataFrames on the 'Date' column to get the final DataFrame for new data
merged_df_new = reduce(lambda left, right: pd.merge(left, right, on='Date', how='outer'), new_strategy_data_frames.values())

# Combine the existing DataFrame with the new merged DataFrame
merged_df_combined = existing_df.copy()

# Iterate through the new ticker names and update prices if the ticker already exists, otherwise add the new column
for ticker in ticker_names_new:
    if ticker in ticker_names_existing:
        merged_df_combined[ticker] = merged_df_new[ticker]
    else:
        merged_df_combined = pd.merge(merged_df_combined, merged_df_new[['Date', ticker]], on='Date', how='left')


# Fill the blanks with NaN
merged_df_combined = merged_df_combined.fillna(pd.NA)
# Open the existing Excel file with openpyxl
book = load_workbook(existing_file)

# Select the existing sheet
sheet = book[existing_sheet_name]

# Calculate the start row (2nd row in this case)
start_row = 2

# Convert the DataFrame to a list of rows
rows = dataframe_to_rows(merged_df_combined, index=False, header=True)

# Append the rows to the existing sheet starting from the 2nd row
for r_idx, row in enumerate(rows, start_row):
    for c_idx, value in enumerate(row, 1):
        sheet.cell(row=r_idx, column=c_idx, value=value)

# Save the updated Excel file
book.save(existing_file)
print("Data added to the existing sheet successfully!")
# Save the combined DataFrame to a new Excel file
#output_file = 'combined_strategies.xlsx'
#merged_df_combined.to_excel(existing_file, existing_sheet_name , index=False, startrow = 1)

