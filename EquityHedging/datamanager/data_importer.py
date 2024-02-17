# -*- coding: utf-8 -*-
"""
Created on Fri Jul 14 2023

@author: Powis Forjoe
"""

import pandas as pd
import xlrd
from openpyxl import load_workbook

from . import data_manager_new as dm


class ExcelImporter:
    def __init__(self, file_path, sheet_name=None, index_col=0, skip_rows=None, use_cols=None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.index_col = index_col
        if skip_rows is None:
            skip_rows = []
        self.skip_rows = skip_rows
        self.use_cols = use_cols

    def read_excel_data(self):
        """
        Reads an Excel file into a dataframe or dictionary

        Returns
        -------
        dataframe/dictionary of dataframes

        """

        data = pd.read_excel(self.file_path, sheet_name=self.sheet_name, index_col=self.index_col,
                             skiprows=self.skip_rows, usecols=self.use_cols)
        return self.get_real_cols(data)

    def get_excel_sheet_names(self):
        sheets = []
        if self.file_path.endswith('.xlsx'):
            workbook = load_workbook(filename=self.file_path, read_only=True)
            sheets = workbook.sheetnames
        elif self.file_path.endswith('.xls'):
            workbook = xlrd.open_workbook(self.file_path)
            sheets = workbook.sheet_names()
        return sheets

    @staticmethod
    def get_real_cols(data):
        """
        Removes empty columns labeled 'Unnamed: ' after importing excel data

        Parameters:
        df -- dataframe

        Returns:
        dataframe
        """
        if isinstance(data, dict):
            for key in data:
                real_cols = [x for x in data[key].columns if not x.startswith("Unnamed: ")]
                data[key] = data[key][real_cols]
        else:
            real_cols = [x for x in data.columns if not x.startswith("Unnamed: ")]
            data = data[real_cols]
        return data


# TODO: add skip_cols
class DataImporter:

    def __init__(self, file_path, sheet_name=0, index_col=0, skip_rows=None,
                 data_source='custom', col_dict=None, drop_na=False, index_data=False):
        """
        Reads an Excel file into a dataImporter object

        Parameters
        ----------
        file_path : string
            Valid string path.
        sheet_name : int, string, list of strings, ints, optional
            name(s) of Excel sheet(s) or positions. The default is 0.
        index_col : int, optional
            The default is 0.
        skip_rows : list, optional
            list of rows to skip when importing. The default is [].
        data_source : string, optional
            source of Excel file. The default is 'custom'.
        col_dict :
        drop_na : bool, optional
            drop NaN values. The default is True.
        index_data : TYPE, optional
            DESCRIPTION. The default is False.

        Returns
        -------
        dataImporter object

        """
        if col_dict is None:
            col_dict = {}
        if skip_rows is None:
            skip_rows = []
        self.file_path = file_path
        self.sheet_name = sheet_name  # if sheet_name != 0 else get_excel_sheet_names(self.file_path)
        self.index_col = index_col
        self.skip_rows = skip_rows
        self.data_source = data_source
        self.col_dict = col_dict
        self.drop_na = drop_na
        self.index_data = index_data
        self.data_import = self.import_data()
        self.data_dict_bool = isinstance(self.data_import, dict)

    def import_data(self):
        """
        import data from an Excel file into a Dataframe or Dictionary

        Returns
        -------
        dataframe or Dictionary of dataframes

        """
        data = ExcelImporter(file_path=self.file_path, sheet_name=self.sheet_name, index_col=self.index_col,
                             skip_rows=self.skip_rows).read_excel_data()
        # drop nas
        if self.drop_na:
            data = dm.drop_nas(data)

        # rename columns
        if bool(self.col_dict):
            data = dm.rename_columns(data, self.col_dict)

        if isinstance(data, dict):
            self.data_dict_bool = True
            self.sheet_name = list(data.keys())

        return data


class InnocapDataImporter(DataImporter):
    def __init__(self, file_path, sheet_name=0, index_col=None, skip_rows=None,
                 data_source='innocap', col_dict=None, drop_na=False, index_data=False):
        """
        Reads an Excel file into an innocapDataImporter object

        Parameters
        ----------
        file_path : string
            Valid string path.
        sheet_name : int, string, list of strings, ints, optional
            name(s) of Excel sheet(s) or positions. The default is 0.
        index_col : int, optional
            The default is None.
        skip_rows : list, optional
            list of rows to skip when importing. The default is [0,1].
        data_source : string, optional
            source of Excel file. The default is 'innocap'.
        col_dict :
        drop_na : bool, optional
            drop NaN values. The default is False.
        index_data : bool, optional
            is the data index data or return data. The default is False.

        Returns
        -------
        innocapDataImporter object

        """
        if col_dict is None:
            col_dict = {}
        if skip_rows is None:
            skip_rows = [0, 1]
        super().__init__(file_path=file_path, sheet_name=sheet_name, index_col=index_col, skip_rows=skip_rows,
                         data_source=data_source, col_dict=col_dict, drop_na=drop_na, index_data=index_data)


class BbgDataImporter(InnocapDataImporter):

    def __init__(self, file_path, sheet_name=0, index_col=0, skip_rows=None, drop_na=False,
                 index_data=True):
        """
        Reads an Excel file into an bbgDataImporter object

        Parameters
        ----------
        file_path : string
            Valid string path.
        sheet_name : int, string, list of strings, ints, optional
            name(s) of Excel sheet(s) or positions. The default is 0.
        index_col : int, optional
            The default is 0.
        skip_rows : list, optional
            list of rows to skip when importing. The default is [0,1,2,4,5,6].
        drop_na : bool, optional
            drop NaN values. The default is False.
        index_data : TYPE, optional
            DESCRIPTION. The default is True.

        Returns
        -------
        bbgDataImporter object

        """
        if skip_rows is None:
            skip_rows = [0, 1, 2, 4, 5, 6]
        self.file_path = file_path
        self.col_dict = self.get_col_dict()

        super().__init__(file_path=self.file_path, sheet_name=sheet_name, index_col=index_col, skip_rows=skip_rows,
                         data_source='bbg', col_dict=self.col_dict, drop_na=drop_na, index_data=index_data)

        # rename index col
        if isinstance(self.data_import, dict):
            for keys in self.data_import:
                self.data_import[keys].index.names = ['Dates']
        else:
            self.data_import.index.names = ['Dates']

    def get_col_dict(self):
        try:
            keys_df = ExcelImporter(file_path=self.file_path, sheet_name='key', index_col=None,
                                    use_cols='A:B').read_excel_data()
            return keys_df.set_index('Index').to_dict()['Description']
        except Exception as e:
            print(e)
            return {}


class NexenDataImporter(DataImporter):
    def __init__(self, file_path, sheet_name=0, index_col=None, skip_rows=[],
                 col_dict={}, drop_na=False, index_data=False):
        """
        Reads an Excel file into a nexenDataImporter object

        Parameters
        ----------
        file_path : string
            Valid string path.
        sheet_name : int, string, list of strings, ints, optional
            name(s) of Excel sheet(s) or positions. The default is 0.
        index_col : int, optional
            The default is None.
        skip_rows : list, optional
            list of rows to skip when importing. The default is []
        drop_na : bool, optional
            drop NaN values. The default is False.
        index_data : TYPE, optional
            DESCRIPTION. The default is False.

        Returns
        -------
        nexenDataImporter object

        """
        super().__init__(file_path=file_path, sheet_name=sheet_name, index_col=index_col, skip_rows=skip_rows,
                         data_source='nexen', col_dict=col_dict, drop_na=drop_na, index_data=index_data)

        self.data_import = self.data_import[self.col_dict.values()] if bool(self.col_dict) else self.data_import

# class PutSpreadDataImporter(DataImporter):
#     def __init__(self, file_path ,sheet_name = "Daily", index_col=0, skip_rows=[1],
#                  data_source='custom', drop_na=False, index_data = False):
#         """
#         Reads an Excel file into a PutSpreadDataImporter object

#         Parameters
#         ----------
#         file_path : string
#             Valid string path.
#         sheet_name : int, string, list of strings, ints, optional
#             name(s) of Excel sheet(s) or positions. The default is 0.
#         index_col : int, optional
#             The default is None.
#         skip_rows : list, optional
#             list of rows to skip when importing. The default is [].
#         data_source : string, optional
#             source of Excel file. The default is 'nexen'.
#         drop_na : bool, optional
#             drop NaN values. The default is False.
#         index_data : TYPE, optional
#             DESCRIPTION. The default is False.

#         Returns
#         -------
#         PutSpreadDataImporter object

#         """
#         super().__init__(file_path, sheet_name, index_col, skip_rows,
#                               data_source,drop_na, index_data)

#         self.data_import.columns = ['99 Rep','Short Put','Put Spread']


# def get_qis_uni_dict():
#     qis_uni = {}
#     sheet_names = DataImporter.get_excel_sheet_names(dl.QIS_UNIVERSE + "QIS Universe Time Series TEST.xlsx")
#     for sheet in sheet_names:
#         index_price = pd.read_excel(dl.QIS_UNIVERSE + "QIS Universe Time Series TEST.xlsx", sheet_name=sheet,
#                                     index_col=0,
#                                     header=1)
#         qis_uni[sheet] = dm.format_data(index_price, freq='W')
#     return qis_uni
