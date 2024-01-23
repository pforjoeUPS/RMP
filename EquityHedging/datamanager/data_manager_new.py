# -*- coding: utf-8 -*-
"""
Created on Tue Oct  1 17:59:28 2019

@author: Powis Forjoe, Maddie Choi
"""

import os
from datetime import datetime as dt
from math import prod

import numpy as np
import pandas as pd
from openpyxl import load_workbook

CWD = os.getcwd()
DATA_FP = CWD + '\\EquityHedging\\data\\'
RETURNS_DATA_FP = DATA_FP + 'returns_data\\'
EQUITY_HEDGING_RETURNS_DATA = RETURNS_DATA_FP + 'eq_hedge_returns.xlsx'
# RETURNS_DATA_FP = CWD +'\\EquityHedging\\data\\'
# EQUITY_HEDGING_RETURNS_DATA = DATA_FP + 'ups_equity_hedge\\returns_data.xlsx'
NEW_DATA = DATA_FP + 'new_strats\\'
# UPDATE_DATA = RETURNS_DATA_FP + 'update_strats\\'
UPDATE_DATA = DATA_FP + 'update_data\\'
EQUITY_HEDGE_DATA = DATA_FP + 'ups_equity_hedge\\'

QIS_UNIVERSE = CWD + '\\Cluster Analysis\\data\\'

NEW_DATA_COL_LIST = ['SPTR', 'SX5T', 'M1WD', 'Long Corp', 'STRIPS', 'Down Var',
                     'Vortex', 'VOLA I', 'VOLA II', 'Dynamic VOLA', 'Dynamic Put Spread',
                     'GW Dispersion', 'Corr Hedge', 'Def Var (Mon)', 'Def Var (Fri)', 'Def Var (Wed)',
                     'Commodity Basket']

LIQ_ALTS_MGR_DICT = {'Global Macro': ['1907 Penso Class A', 'Bridgewater Alpha', 'DE Shaw Oculus Fund',
                                      'Element Capital'],
                     'Trend Following': ['1907 ARP TF', '1907 Campbell TF', '1907 Systematica TF',
                                         'One River Trend'],
                     'Absolute Return': ['1907 ARP EM', '1907 III CV', '1907 III Class A',
                                         'ABC Reversion', 'Acadian Commodity AR',
                                         'Blueshift', 'Duality', 'Elliott'],
                     }


def merge_dicts(main_dict, new_dict, drop_na=True, fillzeros=False):
    """
    Merge new_dict to main_dict

    Parameters:
    main_dict -- dictionary
    new_dict -- dictionary
    drop_na -- bool
    fillzeros -- bool

    Returns:
    dictionary
    """

    merged_dict = {}
    for key in main_dict:
        try:
            df_main = main_dict[key].copy()
            df_new = new_dict[key]
            if key == 'Daily':
                df_main = df_main.fillna(0)
                merged_dict[key] = merge_data_frames(df_main, df_new, drop_na=drop_na, fillzeros=True)

            else:
                merged_dict[key] = merge_data_frames(df_main, df_new, drop_na=drop_na, fillzeros=fillzeros)
        except KeyError:
            pass
    return merged_dict


def merge_data_frames(df_main, df_new, drop_na=True, fillzeros=False, how='outer'):
    """
    Merge df_new to df_main and drop na values

    Parameters:
    df_main -- dataframe
    df_new -- dataframe
    drop_na -- bool
    fillzeros -- bool

    Returns:
    dataframe
    """

    df = pd.merge(df_main, df_new, left_index=True, right_index=True, how=how)
    if drop_na:
        df.dropna(inplace=True)
    if fillzeros:
        df = df.fillna(0)
    return df


def get_columns_data(data):
    if isinstance(data, dict):
        col_data = {}
        for key in data:
            col_data[key] = list(data[key].columns)
    else:
        return list(data.columns)


# TODO: move to dxf
def resample_data(df, freq="M"):
    data = df.copy()
    data.index = pd.to_datetime(data.index)
    if not (freq == 'D'):
        data = data.resample(freq).ffill()
    return data


# TODO: move to dxf
def format_data(df_index, freq="M", dropna=True, drop_zero=False):
    """
    Format dataframe, by freq, to return dataframe

    Parameters:
    df_index -- dataframe
    freq -- string ('M', 'W', 'D')

    Returns:
    dataframe
    """
    data = resample_data(df_index, freq)
    data = data.pct_change(1)
    data = data.iloc[1:, ]
    if dropna:
        data.dropna(inplace=True)

    if drop_zero:
        data = data.loc[(data != 0).any(1)]
    return data


def get_min_max_dates(df_returns):
    """
    Return a dict with the min and max dates of a dataframe.

    Parameters:
    df_returns -- dataframe with index dates

    Returns:
    dict(key (string), value(dates))
    """
    # get list of date index
    dates_list = list(df_returns.index.values)
    dates = {}
    dates['start'] = dt.utcfromtimestamp(dates_list[0].astype('O') / 1e9)
    dates['end'] = dt.utcfromtimestamp(dates_list[len(dates_list) - 1].astype('O') / 1e9)
    return dates


def compute_no_of_years(df_returns):
    """
    Returns number of years in a dataframe based off the min and max dates

    Parameters:
    df_returns -- dataframe with returns

    Returns:
    double
    """

    min_max_dates = get_min_max_dates(df_returns)
    no_of_years = (min_max_dates['end'] - min_max_dates['start']).days / 365
    return no_of_years


def switch_freq_int(arg):
    """
    Return an integer equivalent to frequency in years

    Parameters:
    arg -- string ('D', 'W', 'M')

    Returns:
    int of frequency in years
    """
    switcher = {
        "D": 252,
        "B": 252,
        "W": 52,
        "M": 12,
        "Q": 4,
        "Y": 1,
        "A": 1,
        "1D": 252,
        "1W": 52,
        "1M": 12,
        "1Q": 4,
        "1Y": 1,
        "1A": 1,
    }
    return switcher.get(arg, 252)


def switch_freq_list(arg):
    """
    Return an integer equivalent to frequency in years

    Parameters:
    arg -- string ('D', 'W', 'M')

    Returns:
    int of frequency in years
    """
    switcher = {
        "D": ['D', 'W', 'M', 'Q', 'Y'],
        "W": ['W', 'M', 'Q', 'Y'],
        "M": ['M', 'Q', 'Y'],
        "Q": ['Q', 'Y'],
        "Y": ['Y'],
    }
    return switcher.get(arg, ['D', 'W', 'M', 'Q', 'Y'])


def switch_freq_string(arg):
    """
    Return an string equivalent to frequency
    eg: swith_freq_string('D') returns 'Daily'

    Parameters:
    arg -- string ('D', 'W', 'M')

    Returns:
    string
    """
    switcher = {
        "D": "Daily",
        "B": "Daily",
        "W": "Weekly",
        "M": "Monthly",
        "Q": "Quarterly",
        "Y": "Yearly",
        "A": "Yearly",
        "1D": "Daily",
        "1W": "Weekly",
        "1M": "Monthly",
        "1Q": "Quarterly",
        "1Y": "Yearly",
        "1A": "Yearly",
    }
    return switcher.get(arg, 'Daily')


def switch_string_freq(arg):
    """
    Return an string equivalent to freq
    eg: swith_freq_string('Daily') returns 'D'

    Parameters:
    arg -- string ('D', 'W', 'M')

    Returns:
    string
    """
    switcher = {
        "Daily": "D",
        "Weekly": "W",
        "Monthly": "M",
        "Quarterly": "Q",
        "Yearly": "Y",
    }
    return switcher.get(arg, 'D')


def get_freq(returns_df):
    dates = returns_df.index
    date_rng = pd.to_datetime(dates)  # DatetimeIndex
    freq = pd.infer_freq(date_rng)
    if freq is not None:
        return freq[0]
    else:
        return 'D'


def get_freq_data(returns_df):
    freq = get_freq(returns_df)
    return {'freq': freq, 'freq_string': switch_freq_string(freq), 'freq_int': switch_freq_int(freq)}


def remove_na(df, col_name):
    """
    Remove na values from column

    Parameters:
    df -- dataframe
    col_name -- string (column name in dataframe)

    Returns:
    dataframe
    """
    clean_df = pd.DataFrame(df[col_name].copy())
    clean_df.dropna(inplace=True)
    return clean_df


def get_freq_ratio(freq1, freq2):
    """
    Returns ratio of 2 frequencies

    Parameters:
    freq1 -- string
    freq2 -- string

    Returns:
    int
    """

    return round(switch_freq_int(freq1) / switch_freq_int(freq2))


def convert_to_freq2(arg, freq1, freq2):
    """
    Converts number from freq1 to freq2

    Parameters:
    arg -- int
    freq1 -- string
    freq2 -- string

    Returns:
    int
    """

    return round(arg / get_freq_ratio(freq1, freq2))


# TODO: no longer needed dh does this now
def get_vrr_weights(weights):
    """
    Returns VRR weights from notional weights

    Parameters:
    notional weights -- list

    Returns:
    list
    """
    notional_vrr_weights = [weights[4], weights[5]]
    port_total = float(sum(notional_vrr_weights))
    vrr_weights = [weight / port_total for weight in notional_vrr_weights]
    return vrr_weights


# TODO: no longer needed dh does this now
def create_vrr_portfolio(returns, weights):
    """
    Updates returns to combine VRR2 and VRRTrend returns into VRRPortfolio

    Parameters:
    df_returns -- dataframe
    weights -- list

    Returns:
    dataframe
    """
    returns_dict = returns.copy()
    vrr_weights = get_vrr_weights(weights)
    freqs = ['Daily', 'Weekly', 'Monthly', 'Quarterly', 'Yearly']
    for freq in freqs:
        vrr_portfolio = returns_dict[freq]['VRR 2'] * vrr_weights[0] + returns_dict[freq]['VRR Trend'] * vrr_weights[1]
        returns_dict[freq].insert(loc=4, column='VRR Portfolio', value=vrr_portfolio)
        returns_dict[freq].drop(['VRR Trend'], inplace=True, axis=1)
        returns_dict[freq].drop(['VRR 2'], inplace=True, axis=1)
    return returns_dict


def drop_nas(data):
    if isinstance(data, dict):
        for key in data:
            data[key].dropna(inplace=True)
    else:
        data.dropna(inplace=True)
    return data


def check_col_len(df, col_list):
    if len(col_list) != len(df.columns):
        return list(df.columns)
    else:
        return col_list


def rename_columns(data, col_dict):
    if isinstance(data, dict):
        for key in data:
            data[key].rename(columns=col_dict, inplace=True)
    else:
        data.rename(columns=col_dict, inplace=True)
    return data


def get_notional_weights(df_returns):
    """
    Returns list of notional values for stratgies

    Parameters:
    df_returns -- dataframe

    Returns:
    list
    """
    weights = [float(input('notional value (Billions) for ' + col + ': ')) for col in df_returns.columns]
    # df_returns = create_vrr_portfolio(df_returns, weights)
    # weights.append(weights[4]+weights[5])
    # del weights[4:6]
    return weights


# TODO: no longer needed dh does this now
def create_copy_with_fi(df_returns, equity='SPTR', freq='M', include_fi=False):
    """
    Combine columns of df_returns together to get:
    FI Benchmark (avg of Long Corps and STRIPS)
    VOLA (avg of VOLA I and VOLA II)
    Def Var (weighted avg Def Var (Fri): 60%, Def Var (Mon):20%, Def Var (Wed): 20%)

    Parameters:
    df_returns -- dataframe
    freq -- string

    Returns:
    dataframe
    """
    strategy_returns = df_returns.copy()

    strategy_returns['VOLA 3'] = strategy_returns['Dynamic VOLA']
    strategy_returns['Def Var'] = strategy_returns['Def Var (Fri)'] * .4 + strategy_returns['Def Var (Mon)'] * .3 + \
                                  strategy_returns['Def Var (Wed)'] * .3

    if freq == 'W' or freq == 'M':
        if include_fi:
            strategy_returns['FI Benchmark'] = (strategy_returns['Long Corp'] + strategy_returns['STRIPS']) / 2
            strategy_returns = strategy_returns[[equity, 'FI Benchmark',
                                                 # '99%/90% Put Spread',
                                                 'Down Var', 'Vortex', 'VOLA 3', 'Dynamic Put Spread',
                                                 'VRR 2', 'VRR Trend', 'GW Dispersion', 'Corr Hedge', 'Def Var',
                                                 'Commodity Basket']]
        else:
            strategy_returns = strategy_returns[[equity,
                                                 # '99%/90% Put Spread',
                                                 'Down Var', 'Vortex', 'VOLA 3', 'Dynamic Put Spread',
                                                 'VRR 2', 'VRR Trend', 'GW Dispersion', 'Corr Hedge', 'Def Var',
                                                 'Commodity Basket']]
    else:
        strategy_returns = strategy_returns[[equity,
                                             # '99%/90% Put Spread',
                                             'Down Var', 'Vortex',
                                             'VOLA 3', 'Dynamic Put Spread', 'VRR 2', 'VRR Trend',
                                             'GW Dispersion', 'Corr Hedge', 'Def Var', 'Commodity Basket']]

    return strategy_returns


def get_real_cols(df):
    """
    Removes empty columns labeled 'Unnamed: ' after importing data

    Parameters:
    df -- dataframe

    Returns:
    dataframe
    """
    real_cols = [x for x in df.columns if not x.startswith("Unnamed: ")]
    df = df[real_cols]
    return df


def get_equity_hedge_returns(equity='SPTR', include_fi=False, strat_drop_list=[], only_equity=False, all_data=False):
    """
    Returns a dictionary of dataframes containing returns data of
    different frequencies

    Parameters:
    equity -- dataframe
    include_fi --boolean
    strat_drop_list -- list
    only_equity -- boolean

    Returns:
    dictionary
    """
    returns_dict = {}
    freqs = ['D', 'W', 'M', 'Q', 'Y']
    for freq in freqs:
        freq_string = switch_freq_string(freq)
        temp_ret = pd.read_excel(EQUITY_HEDGING_RETURNS_DATA,
                                 sheet_name=freq_string,
                                 index_col=0)
        temp_ret = get_real_cols(temp_ret)
        if all_data:
            returns_dict[freq_string] = temp_ret.copy()
        else:
            returns_dict[freq_string] = create_copy_with_fi(temp_ret, equity, freq, include_fi)
        if strat_drop_list:
            returns_dict[freq_string].drop(strat_drop_list, axis=1, inplace=True)
        if only_equity:
            returns_dict[freq_string] = returns_dict[freq_string][[equity]]
        returns_dict[freq_string].index.names = ['Date']

    return returns_dict


# TODO: move to dxf
def get_data_dict(data, return_data=False, dropna=True):
    """
    Converts daily data into a dictionary of dataframes containing returns
    data of different frequencies

    Parameters:
    data -- df
    data_type -- string

    Returns:
    dictionary
    """
    freq_data = get_freq(data)
    freq_list = switch_freq_list(freq_data)
    data_dict = {}
    # if data_type != 'index':
    if return_data:
        try:
            data.index = pd.to_datetime(data.index)
        except TypeError:
            pass
        data = get_prices_df(data)
    for freq in freq_list:
        freq_string = switch_freq_string(freq)
        data_dict[freq_string] = format_data(data, freq, dropna)
    return data_dict


# TODO: move to dxf
def get_prices_df(df_returns, multiplier=100):
    """"
    Converts returns dataframe to index level dataframe

    Parameters:
    df_returns -- returns dataframe

    Returns:
    index price level - dataframe
    """

    df_index = multiplier * (1 + df_returns).cumprod()

    return update_df_index(df_index, multiplier)


# TODO: move to dxf
def update_df_index(df_index, multiplier=100):
    # insert extra row at top for first month of 100
    data = []
    data.insert(0, {})
    df_prices = pd.concat([pd.DataFrame(data), df_index])

    # fill columns with 100 for row 1
    for col in df_prices.columns:
        df_prices[col][0] = multiplier

    # update index to prior month
    df_prices.index.names = ['Dates']
    df_prices.reset_index(inplace=True)
    pd.set_option('mode.chained_assignment', None)
    df_prices.loc[:, ('Dates')][0] = df_index.index[0] - pd.DateOffset(months=1)
    df_prices.set_index('Dates', inplace=True)

    return df_prices


# TODO: move to dxf
def get_price_series(return_series, multiplier=100):
    df_index = multiplier * (1 + return_series).cumprod()

    return update_df_index(df_index, multiplier)[0]


# TODO: check with script if needed
def get_new_strategy_returns_data(filename, sheet_name, return_data=True, strategy_list=[]):
    """
    dataframe of stratgy returns

    Parameters:
    filename -- string
    sheet_name -- string
    strategy_list -- list

    Returns:
    dataframe
    """
    df_strategy = pd.read_excel(NEW_DATA + filename, sheet_name=sheet_name, index_col=0)
    df_strategy = get_real_cols(df_strategy)
    if strategy_list:
        df_strategy.columns = strategy_list
    try:
        df_strategy.index = pd.to_datetime(df_strategy.index)
    except TypeError:
        pass
    df_strategy = df_strategy.resample('D').ffill()
    new_strategy_returns = df_strategy.copy()
    if return_data is False:
        new_strategy_returns = df_strategy.pct_change(1)
    new_strategy_returns.dropna(inplace=True)
    return new_strategy_returns


def merge_dicts_list(dict_list, drop_na=True):
    '''
    merge main dictionary with a dictionary list

    Parameters
    ----------
    dict_list : list

    Returns
    -------
    main_dict : dictionary
        new dictionary created upon being merged with a list

    '''
    main_dict = dict_list[0]
    dict_list.remove(main_dict)
    # iterate through dictionary
    for new_dict in dict_list:
        # merge each dictionary in the list of dictionaries to the main
        main_dict = merge_dicts(main_dict, new_dict, drop_na)
    return main_dict


# def match_dict_columns(main_dict, new_dict):
#     '''


#     Parameters
#     ----------
#     main_dict : dictionary
#     original dictionary
#     new_dict : dictionary
#     dictionary that needs to have columns matched to main_dict
#     Returns
#     -------
#     new_dict : dictionary
#         dictionary with matched columns

#     '''

#     #iterate through keys in dictionary
#     for key in new_dict:

#         #set column in the new dictionary equal to that of the main
#         new_dict[key] = new_dict[key][list(main_dict[key].columns)]
#     return new_dict

# def append_dict(main_dict, new_dict):
#     '''
#     update an original dictionary by adding information from a new one

#     Parameters
#     ----------
#     main_dict : dictionary
#     new_dict : dictionary

#     Returns
#     -------
#     main_dict : dictionary

#     '''
#     #iterate through keys in dictionary
#     for key in new_dict:

#         #add value from new_dict to main_dict
#         main_dict[key] = main_dict[key].append(new_dict[key])
#     return main_dict

# TODO: Move to analytics
def compound_ret_from_monthly(strat_monthly_returns, strategy):
    monthly_ret = strat_monthly_returns.copy()
    monthly_ret["Year"] = monthly_ret.index.get_level_values('year')

    years = np.unique(monthly_ret["Year"])
    yr_ret = []
    itd_ret = []
    for i in range(0, len(years)):
        # isolate monthly returns for single year
        monthly_ret_by_yr = monthly_ret.loc[monthly_ret.Year == years[i]][strategy]
        # calculate compound return
        comp_ret = prod(1 + monthly_ret_by_yr) - 1
        yr_ret.append(comp_ret)
        itd_ret.append(np.prod((np.array(yr_ret) + 1).tolist()) - 1)

    ret_dict = {"YTD": yr_ret, "ITD": itd_ret}
    ret_df = pd.DataFrame(ret_dict, index=list(years))
    return ret_df


# TODO: Move to analytics
def month_ret_table(returns_df, strategy):
    '''

    Parameters
    ----------
    returns_df : Data Frame

    strategy : String
        Strategy name

    Returns
    -------
    Data Frame

    '''
    # pull monthly returns from dictionary
    month_ret = pd.DataFrame(returns_df[strategy])
    month_ret.dropna(inplace=True)

    # create monthly return data frame with index of years
    month_ret['year'] = month_ret.index.year
    month_ret['month'] = month_ret.index.month_name().str[:3]

    # change monthly returns into a table with x axis as months and y axis as years
    strat_monthly_returns = month_ret.groupby(['year', 'month']).sum()
    yr_itd_ret = compound_ret_from_monthly(strat_monthly_returns, strategy)

    month_table = strat_monthly_returns.unstack()

    # drop first row index
    month_table = month_table.droplevel(level=0, axis=1)

    # re order columns
    month_table = month_table[["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]]

    # Join yearly returns to the monthly returns table
    table = pd.concat([month_table, yr_itd_ret], axis=1)
    table.index.names = [strategy]

    return table


# def transform_nexen_data(filename = 'liq_alts\\Historical Asset Class Returns.xls', return_data = True, fillna = False):
#     nexen_df = pd.read_excel(DATA_FP + filename)
#     nexen_df = nexen_df[['Account Name\n', 'Account Id\n', 'Return Type\n', 'As Of Date\n',
#                            'Market Value\n', 'Account Monthly Return\n']]
#     nexen_df.columns = ['Name', 'Account Id', 'Return Type', 'Date', 'Market Value', 'Return']
#     if return_data:
#         nexen_df = nexen_df.pivot_table(values='Return', index='Date', columns='Name')
#         nexen_df /=100
#     else:
#         nexen_df = nexen_df.pivot_table(values='Market Value', index='Date', columns='Name')
#     if fillna:
#         nexen_df = nexen_df.fillna(0)
#     return nexen_df

# def transform_nexen_data_1(filepath, fillna = False):
#     nexen_df = pd.read_excel(filepath)
#     nexen_df = nexen_df[['Account Name\n', 'Account Id\n', 'Return Type\n', 'As Of Date\n',
#                            'Market Value\n', 'Account Monthly Return\n']]
#     nexen_df.columns = ['Name', 'Account Id', 'Return Type', 'Date', 'Market Value', 'Return']
#     returns_df = nexen_df.pivot_table(values='Return', index='Date', columns='Name')
#     returns_df /=100
#     mv_df = nexen_df.pivot_table(values='Market Value', index='Date', columns='Name')
#     if fillna:
#         return {'returns': returns_df.fillna(0), 'mv': mv_df.fillna(0)}
#     else:
#         return {'returns': returns_df, 'mv': mv_df}

# def transform_bbg_data(filepath, sheet_name='data', col_list=[]):
#     bbg_df = pd.read_excel(filepath, sheet_name=sheet_name,
#                            index_col=0,skiprows=[0,1,2,4,5,6])
#     bbg_df.index.names = ['Dates']
#     if col_list:
#         bbg_df.columns = col_list
#     return bbg_df

# def get_liq_alts_bmks(equity = 'M1WD',include_fi=True):
#     bmks_index = transform_bbg_data(DATA_FP+'liq_alts\\liq_alts_bmks.xlsx')
#     bmks_index = bmks_index[['HFRXM Index','NEIXCTAT Index', 'HFRXAR Index']]
#     bmks_index.columns = ['HFRX Macro/CTA Index', 'SG Trend Index', 'HFRX Absolute Return Index']
#     bmks_ret = format_data(bmks_index)
#     bmks_ret['Liquid Alts Bmk'] = 0.5*bmks_ret['HFRX Macro/CTA Index'] + 0.3*bmks_ret['HFRX Absolute Return Index'] + 0.2*bmks_ret['SG Trend Index']
#     beta_m =  get_equity_hedge_returns(equity=equity, include_fi=include_fi)
#     return merge_data_frames(beta_m['Monthly'][[equity, 'FI Benchmark']],bmks_ret, drop_na=False)

# def get_liq_alts_dict(filename = 'liq_alts\\Monthly Returns Liquid Alts.xls'):
#     liq_alts_ret = transform_nexen_data(filename)
#     liq_alts_mv = transform_nexen_data(filename, False)
#     liq_alts_dict = {}
#     total_ret = pd.DataFrame(index = liq_alts_ret.index)
#     total_mv = pd.DataFrame(index = liq_alts_mv.index)
#     for key in LIQ_ALTS_MGR_DICT:
#         temp_dict = {}
#         temp_ret = liq_alts_ret[LIQ_ALTS_MGR_DICT[key]].copy()
#         temp_mv = liq_alts_mv[LIQ_ALTS_MGR_DICT[key]].copy()
#         temp_dict = get_agg_data(temp_ret, temp_mv, key)
#         if key == 'Trend Following':
#             temp_dict = {'returns': liq_alts_ret[get_sub_mgrs(key)],
#                               'mv':liq_alts_mv[get_sub_mgrs(key)]}
#         total_ret = merge_data_frames(total_ret, temp_ret[[key]])
#         total_mv = merge_data_frames(total_mv, temp_mv[[key]])
#         liq_alts_dict[key] = temp_dict
#     liq_alts_dict['Total Liquid Alts'] = get_agg_data(total_ret, total_mv, 'Total Liquid Alts')
#     return liq_alts_dict

# def get_liq_alts_returns(equity='M1WD', include_fi=True):
#     liq_alts_bmks = get_liq_alts_bmks()
#     liq_alts_port = get_liq_alts_port(get_liq_alts_dict())
#     return merge_data_frames(liq_alts_port,liq_alts_bmks)

# def get_liq_alts_port(liq_alts_dict):
#     liq_alts_port = pd.DataFrame()
#     for key in LIQ_ALTS_MGR_DICT:
#         liq_alts_port = merge_data_frames(liq_alts_port, liq_alts_dict[key]['returns'], drop_na=False)

#     liq_alts_port = merge_data_frames(liq_alts_port, liq_alts_dict['Total Liquid Alts']['returns'][['Total Liquid Alts']],drop_na=False)
#     return liq_alts_port

# TODO: similar to di.get_excel_sheetnames
def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


# TODO: read the whole work book as a dicitonary then format by key
def get_qis_uni_dict():
    qis_uni = {}
    sheet_names = get_sheetnames_xlsx(QIS_UNIVERSE + "QIS Universe Time Series TEST.xlsx")
    for sheet in sheet_names:
        index_price = pd.read_excel(QIS_UNIVERSE + "QIS Universe Time Series TEST.xlsx", sheet_name=sheet, index_col=0,
                                    header=1)
        qis_uni[sheet] = format_data(index_price, freq='W')
    return qis_uni


def check_notional(df_returns, notional_weights=[]):
    """
    Get notional weights if some weights are missing

    Parameters
    ----------
    df_returns : dataframe
        dataframe of returns
    notional_weights : list, optional
        notional weights of strategies. The default is [].

    Returns
    -------
    notional_weights : list

    """
    # create list of df_returns column names
    col_list = list(df_returns.columns)

    # get notional weights for weighted strategy returns if not accurate
    if len(col_list) != len(notional_weights):
        notional_weights = []
        notional_weights = get_notional_weights(df_returns)

    return notional_weights


def get_wgts(mvs_df, total_col=False, add_total_wgt=False):
    if total_col:
        col_len = mvs_df.shape[1] - 1
        wgts_df = mvs_df.iloc[:, :col_len].divide(mvs_df.iloc[:, :col_len].sum(axis=1), axis='rows')
    else:
        wgts_df = mvs_df.divide(mvs_df.sum(axis=1), axis='rows')

    if add_total_wgt:
        total_name = mvs_df.columns[col_len] if total_col else 'Total'
        wgts_df[total_name] = wgts_df.sum(axis=1)
    return wgts_df


def get_agg_data(df_returns, mvs_df, agg_col):
    agg_ret = df_returns.copy()
    agg_mv = mvs_df.copy()
    wgts = get_wgts(agg_mv)
    agg_ret[agg_col] = (agg_ret * wgts).sum(axis=1)
    agg_mv[agg_col] = agg_mv.sum(axis=1)
    # wgts[agg_col] = wgts.sum(axis=1)
    return {'returns': agg_ret[[agg_col]], 'mv': agg_mv[[agg_col]],
            # 'wgts': wgts[[agg_col]]
            }


# def get_abs_ret():
#     ar_list = ['1907 ARP EM', '1907 III CV', '1907 III Class A', 'ABC Reversion',
#                     'Acadian Commodity AR','Blueshift', 'Duality', 'Elliott']
#     liq_alts_data = get_liq_alts_dict()
#     for key in liq_alts_data:
#         liq_alts_data[key] = liq_alts_data[key][ar_list]
#     wgts = liq_alts_data['mv']/liq_alts_data['mv'].sum(axis=1)
#     liq_alts_data['returns']['Absolute Return-no RP'] = (liq_alts_data['returns']*wgts).sum(axis=1)
#     liq_alts_data['mv']['Absolute Return-no RP'] = liq_alts_data['mv'].sum(axis=1)
#     return liq_alts_data

# def get_mgrs_list(full=True,sub_port= 'Global Macro'):
#     mgr_list = []
#     if full:
#         for key in LIQ_ALTS_MGR_DICT:
#             mgr_list = mgr_list + LIQ_ALTS_MGR_DICT[key]
#             mgr_list.append(key)
#     else:
#         mgr_list = mgr_list + LIQ_ALTS_MGR_DICT[sub_port]
#         mgr_list.append(sub_port)
#     return mgr_list

# def get_sub_mgrs(sub_port = 'Global Macro'):
#     mgr_list = []
#     mgr_list = mgr_list + LIQ_ALTS_MGR_DICT[sub_port]
#     mgr_list.append(sub_port)
#     return mgr_list

# def get_sub_ports():
#     return list(LIQ_ALTS_MGR_DICT.keys())


def get_new_strat_data(filename, sheet_name='data', freq='M', index_data=False):
    new_strat = pd.read_excel(DATA_FP + filename, sheet_name=sheet_name, index_col=0)
    if index_data:
        new_strat = format_data(new_strat, freq)
    return new_strat


def get_period_dict(df_returns, freq='M'):
    obs = len(df_returns)
    m = switch_freq_int(freq)
    if obs <= m:
        return {'Full': df_returns}
    elif obs <= 3 * m:
        return {'Full': df_returns, '1 Year': df_returns.iloc[obs - m:, ]}
    elif obs <= 5 * m:
        return {'Full': df_returns, '3 Year': df_returns.iloc[obs - 3 * m:, ], '1 Year': df_returns.iloc[obs - m:, ]}
    elif obs <= 10 * m:
        return {'Full': df_returns, '5 Year': df_returns.iloc[obs - 5 * m:, ],
                '3 Year': df_returns.iloc[obs - 3 * m:, ], '1 Year': df_returns.iloc[obs - m:, ]}
    else:
        return {'Full': df_returns, '10 Year': df_returns.iloc[obs - 10 * m:, ],
                '5 Year': df_returns.iloc[obs - 5 * m:, ], '3 Year': df_returns.iloc[obs - 3 * m:, ],
                '1 Year': df_returns.iloc[obs - m:, ]}


def get_last_n_values(dict, n=1):
    """Returns the last n values of a dictionary.

    Args:
      dict: The dictionary to get the values from.
      n: The number of values to get.

    Returns:
      A list of the last n values of the dictionary.
    """

    # Get the keys of the dictionary in reverse order.
    keys = list(dict.keys())[::-1]

    # Get the last n values of the dictionary.
    values = [dict[key] for key in keys[:n]]

    return values