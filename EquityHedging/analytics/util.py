# -*- coding: utf-8 -*-
"""
Created on Thu Apr 22 00:02:00 2021

@author: Powis Forjoe and Maddie Choi
"""

import pandas as pd
import numpy as np
from EquityHedging.datamanager import data_manager as dm
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook

from sklearn.linear_model import LinearRegression

def get_pos_neg_df(return_series, pos=True):
    """
    Return dataframe with positive/negative returns of col_name as a condition
    
    Parameters
    ----------
    return_series : series
        returns series.
    pos : boolean, optional
        get positive returns. The default is True.

    Returns
    -------
    dataframe

    """
    
    #make copy of series
    ret = return_series.copy()
    
    #filter index for positive/negative returns
    if pos:
        ret_index = ret.index[ret > 0]
    else:
        ret_index = ret.index[ret < 0]
    
    #create new series
    return ret.loc[ret_index]
    
def get_rolling_cum(df_returns, interval):
    """
    Get rolling cum returns dataframe

    Parameters
    ----------
    df_returns : dataframe
        returns dataframe.
    interval : int

    Returns
    -------
    rolling_cum_ret : dataframe

    """
   
    rolling_cum_ret = df_returns.copy()
    for col in df_returns.columns:
        rolling_cum_ret[col] = (1+rolling_cum_ret[col]).rolling(window=interval).apply(np.prod, raw=True)-1
    rolling_cum_ret.dropna(inplace=True)
    return rolling_cum_ret

def get_strat_weights(notional_weights, include_fi=False):
    """
    Return weights of each strategy compared to Equity or Equity and FI

    Parameters
    ----------
    notional_weights : list, optional
        notional weights of strategies. The default is [].
    include_fi : boolean, optional
        Include Fixed Income benchmark. The default is False.
    
    Returns
    -------
    strat_weights : list

    """

    #get strategy weights (strat_weights)
    strat_total = 0
    for weight in notional_weights:
        strat_total += weight
    strat_total = strat_total - notional_weights[0]
    if(include_fi):
        strat_total = strat_total - notional_weights[1]
    strat_total
    
    strat_weights = [weight / strat_total for weight in notional_weights]
    strat_weights[0] = 0
    if(include_fi):
        strat_weights[1] = 0
    return strat_weights

# def get_weighted_strats_df(df_returns, notional_weights):
#     """
#     Return dataframe of weighted startegy returns, with and without newest strategy

#     Parameters:
#     returns_df -- dataframe
#     notional weights -- list
    
#     Returns:
#     df_weighted_strats -- dataframe
#     """
    
#     #Get weighted strategies (weighted_strats) with and 
#     #without new strategy (weighted_strats_old)
#     pct_weights = [weight / notional_weights[0] for weight in notional_weights]
#     pct_weights
#     pct_weights.append(0)
#     pct_weights
    
#     pct_weights_old = pct_weights.copy()
#     pct_weights_old[len(pct_weights_old)-2]=0
#     pct_weights_old
    
#     df_weighted_strats = df_returns.dot(tuple(pct_weights)).to_frame()
#     df_weighted_strats.columns = ['Weighted Strats']
#     col_names = [col for col in df_returns]
#     wgt_strat_wo_name = 'Weighted Strats w/o ' + col_names[len(col_names)-2]
#     df_weighted_strats[wgt_strat_wo_name] = df_returns.dot(tuple(pct_weights_old)).to_frame()
#     return df_weighted_strats

def get_pct_weights(notional_weights, include_fi=False):
    """
    Return percentage weights based off of notional weights
    
    Parameters
    ----------
    notional_weights : list, optional
        notional weights of strategies. The default is [].
    include_fi : boolean, optional
        Include Fixed Income benchmark. The default is False.
    
    Returns
    -------
    pct_weights : list

    """
    
    total_weight = notional_weights[0] + notional_weights[1] if include_fi else notional_weights[0]
    pct_weights = [weight / total_weight for weight in notional_weights]
    return pct_weights

def get_df_weights(notional_weights, col_list, include_fi=False):
    """
    Returns dataframe with portoflio weighting information

    Parameters
    ----------
    notional_weights : list, optional
        notional weights of strategies. The default is [].
    col_list : string
        list of strategies.
    include_fi : boolean, optional
        Include Fixed Income benchmark. The default is False.

    Returns
    -------
    df_weights : dataframe

    """
    
    #define index of df_weights
    index_list = ['Notional Weights (Billions)',
                  'Percentage Weights',
                  'Strategy Weights']
    

    
    #compute percentage and strategy weights
    pct_weights = get_pct_weights(notional_weights, include_fi)
    strat_weights = get_strat_weights(notional_weights, include_fi)
    
    #create df_weights
    df_weights = pd.DataFrame([notional_weights, pct_weights, strat_weights],
                              index = index_list)
    #rename columns
    df_weights.columns = col_list

    return df_weights

def check_notional(df_returns, notional_weights=[]):
    """
    Get notional weights if some weights are missing

    Parameters
    ----------
    df_returns : dataframe
        dataframe of returns
    notional_weights : list, optional
        notional weights of strategies. The default is [].
    
    Returns
    -------
    notional_weights : list

    """
    #create list of df_returns column names
    col_list = list(df_returns.columns)
    
    #get notional weights for weighted strategy returns if not accurate
    if len(col_list) != len(notional_weights):
        notional_weights = []
        notional_weights = dm.get_notional_weights(df_returns)
    
    return notional_weights

def get_weighted_strats_df(df_returns, notional_weights=[], include_fi=False, new_strat=False,num_new_strats=1):
    """
    Return dataframe of weighted strategy returns, with and without newest strategy

    Parameters
    ----------
    df_returns : dataframe
        dataframe of returns
    notional_weights : list, optional
        notional weights of strategies. The default is [].
    include_fi : boolean, optional
        Include Fixed Income benchmark. The default is False.
    new_strat : boolean, optional
        Does analysis involve a new strategy. The default is False.
    num_new_strats: integer, optional
        Number of new strategies to consider. The default is 1.
    Returns
    -------
    df_weighted_strats : dataframe

    """
    
    #confirm notional weights is correct length
    notional_weights = check_notional(df_returns, notional_weights)
    
    #get weighted strategies
    pct_weights = get_pct_weights(notional_weights, include_fi)
    df_weighted_strats = df_returns.dot(tuple(pct_weights)).to_frame()
    df_weighted_strats.columns = ['Weighted Strats']
    
    #get weighted strategies without new strat
    if new_strat:
        pct_weights_old = pct_weights.copy()
        #pct_weights_old[len(pct_weights_old)-num_new_strats] = 0
        for i in range(num_new_strats):
            pct_weights_old[len(pct_weights_old) - (i+1)] = 0
    

        col_names = list(df_returns.columns)
        wgt_strat_wo_name = 'Weighted Strats w/o New Strategies'
        df_weighted_strats[wgt_strat_wo_name] = df_returns.dot(tuple(pct_weights_old)).to_frame()
    
    return df_weighted_strats

def get_weighted_hedges(df_returns, notional_weights, include_fi=False, new_strat=False, weight_col = 'Weighted Hedges', num_new_strats=1):
    """
    Return dataframe of weighted hedge returns, with and without newest strategy

    Parameters
    ----------
    df_returns : dataframe
        dataframe of returns
    notional_weights : list, optional
        notional weights of strategies. The default is [].
    include_fi : boolean, optional
        Include Fixed Income benchmark. The default is False.
    new_strat : boolean, optional
        Does analysis involve a new strategy. The default is False.
    num_new_strats: integer, optional
        Number of new strategies to consider. The default is 1.
    Returns
    -------
    df_weighted_hedges : dataframe

    """
    
    #get weighted hedges
    df_weighted_hedges = df_returns.copy()
    notional_weights = check_notional(df_weighted_hedges,notional_weights)
    strat_weights = get_strat_weights(notional_weights, include_fi)
    df_weighted_hedges['Weighted Hedges'] = df_weighted_hedges.dot(tuple(strat_weights))
    
    #get weighted hedges w/o new strategy
    if new_strat:
        col_list = list(df_returns.columns)
        temp_weights = notional_weights.copy()
        #temp_weights[len(temp_weights)-num_new_strats] = 0
        for i in range(num_new_strats):
            temp_weights[len(temp_weights) - (i+1)] = 0
        temp_strat_weights = get_strat_weights(temp_weights, include_fi)
        wgt_hedge_wo_name = 'Weighted Hedges w/o New Strategies'
        df_weighted_hedges[wgt_hedge_wo_name] = df_returns.dot(tuple(temp_strat_weights))
        

    return df_weighted_hedges

#TODO: Add comments
def bucket(x):
    """
    """
    
    if x < 1.0:
        return 'Bottom'
    elif x < 2.0 :
        return '2nd'
    elif x < 3.0 :
        return '3rd'
    elif x < 4.0 :
        return '4th'
    else:
        return 'Top'

#TODO: Add comments
def decile_bucket(x):
    if x < 1.0:
        return 'Bottom'
    elif x < 2.0 :
        return '2nd'
    elif x < 3.0 :
        return '3rd'
    elif x < 4.0 :
        return '4th'
    elif x < 5.0 :
        return '5th'
    elif x < 6.0 :
        return '6th'
    elif x < 7.0 :
        return '7th'
    elif x < 8.0 :
        return '8th'
    elif x < 9.0 :
        return '9th'
    else:
        return 'Top'

   
def get_normalized_data(df):
    '''

    Parameters
    ----------
    df : data frame
        data frame of returns data for a given frequency

    Returns
    -------
    df_normal : data frame
        normalizes data to be within the range 0 to 1

    '''
    scaler= MinMaxScaler()
    
    #creates data frame with normalized data
    df_normal = pd.DataFrame(scaler.fit_transform(df), columns = df.columns, index = df.index )

    return df_normal



def convert_dict_to_df(dict,index = []):
    '''

    Parameters
    ----------
    dict : dictionary
        Input a dictionary that will be turned into a data frame. 
    index : list
        Index (aka row) names. The default is [].
     
    ** Note the data frame column names will be the keys in the dictionary **
    
    Returns
    -------
    Data Frame
 
    '''
    
    df=pd.DataFrame(dict, index = index )

    return df

#todo: change mult cols
def reverse_signs_in_col(df, col_name):
    '''
    

    Parameters
    ----------
    df : data frame

    Returns
    -------
    df_reverse : data frame
        same data frame as in the input, but with downside reliability terms as positive.

    '''
    df_reverse = df.copy()
    if col_name in df.columns:
        for x in df_reverse.index:
            df_reverse[col_name][x] = -(df_reverse[col_name][x])
            
    return df_reverse

def change_to_neg(df):
    '''
    

    Parameters
    ----------
    df : data frame

    Returns
    -------
    df_reverse : data frame
        same data frame as in the input, but with downside reliability terms as positive.

    '''
    df_reverse = df.copy()
    for col_name in df.columns:
        for x in df_reverse.index:
            if df_reverse[col_name][x] > 0:
                df_reverse[col_name][x] = -(df_reverse[col_name][x])
            
    return df_reverse

def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def append_dict_dfs(dictionary):
    '''
    

    Parameters
    ----------
    dictionary : dictionary

    Returns
    -------
    Dataframe that appends all dataframes within dictionary into one. 

    '''
    df = pd.DataFrame()
    for i in list(dictionary.keys()):
        temp_df = dictionary[i]
        df = df.append(temp_df)
            
    return df

def reg_helper(x_position, y_position):
    '''
    Perform linear regression analysis and return regression data.

    Parameters:
        x_position (numpy.ndarray): The independent variable data.
        y_position (numpy.ndarray): The dependent variable data.

    Returns:
        list: A list containing the regression data.
            - Index 0: x values for plotting the regression line.
            - Index 1: Predicted y values for the corresponding x values.
            - Index 2: Intercept of the regression line.
            - Index 3: Coefficient of x in the regression equation.
            - Index 4: Beta value (same as the coefficient).

    '''
    reg_data = []  # Create an empty list to store regression data
    
    # Create a linear regression model and fit it with the given data
    regression_model_pos = LinearRegression()
    regression_model_pos.fit(x_position, y_position)
    
    # Generate a range of x values for plotting the regression line
    x_range_pos = np.linspace(min(x_position), max(x_position), 100)
    
    # Predict the corresponding y values for the generated x values
    y_pred_pos = regression_model_pos.predict(x_range_pos.reshape(-1, 1))
    
    # Extract the intercept and coefficient of the regression model
    intercept_pos = regression_model_pos.intercept_
    coefficient_pos = regression_model_pos.coef_[0]
    
    # Calculate the beta (coefficient) value
    beta_pos = coefficient_pos
    
    # Append the regression data to the reg_data list
    reg_data.append(x_range_pos)  # x values for plotting
    reg_data.append(y_pred_pos)   # predicted y values
    reg_data.append(intercept_pos)  # intercept of the regression line
    reg_data.append(coefficient_pos)  # coefficient of x in the regression equation
    reg_data.append(beta_pos)  # beta value
    
    return reg_data


def regression(frequency, returns, strategy_y, strategy_x = 'SPTR'):
    '''
    Perform regression analysis on the given data.

    Parameters:
    frequency (str): The frequency of the returns data.
    returns (pandas.DataFrame): The returns data.
    strategy_y (str): The dependent variable strategy.
    strategy_x (str, optional): The equity benchmark. Defaults to 'SPTR'.

    Returns:
    tuple: A tuple containing the regression data.
        - x_pos (numpy.ndarray): Independent variable values for normal data.
        - y_pos (numpy.ndarray): Dependent variable values for normal data.
        - x_neg (numpy.ndarray): Independent variable values for tail data.
        - y_neg (numpy.ndarray): Dependent variable values for tail data.
        - normal_data (list): Regression data for normal data.
        - tail_data (list): Regression data for tail data.

    '''
    # Set the strategy for comparison with an equity benchmark
    comparison_strategy = strategy_y
    # Copy the returns data for the specified frequency
    data = returns[frequency].copy()

    # Check if strategy_x is 'VIX' or 'UX3'
    if(strategy_x == 'VIX' or strategy_x == 'UX3'):
        # Filter data based on 'SPTR' values below the 2.5th percentile
        data_tail = data[data['SPTR'] < np.quantile(data['SPTR'],.025)]
        # Filter data based on 'SPTR' values above or equal to the 2.5th percentile
        data_normal = data[data['SPTR'] >= np.quantile(data['SPTR'],.025)]
    else:
        # Filter data based on strategy_x values below the 2.5th percentile
        data_tail = data[data[strategy_x] < np.quantile(data[strategy_x],.025)]
        # Filter data based on strategy_x values above or equal to the 2.5th percentile
        data_normal = data[data[strategy_x] >= np.quantile(data[strategy_x],.025)]

    # Prepare data for regression analysis
    x_pos = data_normal[strategy_x].values.reshape(-1, 1)
    y_pos = data_normal[comparison_strategy].values
    x_neg = data_tail[strategy_x].values.reshape(-1, 1)
    y_neg = data_tail[comparison_strategy].values
    x_all = data[strategy_x].values.reshape(-1, 1)
    y_all = data[comparison_strategy].values

    # Perform regression analysis on normal data
    normal_data = reg_helper(x_pos, y_pos)
    # Perform regression analysis on tail data
    tail_data = reg_helper(x_neg, y_neg)
    # Perform regression analysis on all data
    all_data = reg_helper(x_all, y_all)

    # Print regression equations
    print(f"Regression equation (All Data): {comparison_strategy} = {all_data[3]:.4f} * {strategy_x} + {all_data[2]:.4f}")
    print(f"Regression equation ({strategy_x} Highest 97.5%): {comparison_strategy} = {normal_data[3]:.4f} * {strategy_x} + {normal_data[2]:.4f}")
    print(f"Regression equation ({strategy_x} Lowest 2.5%): {comparison_strategy} = {tail_data[3]:.4f} * {strategy_x} + {tail_data[2]:.4f}")

    # Print beta values
    print(f"Beta (All Data): {all_data[4]:.4f}")
    print(f"Beta ({strategy_x} Highest 97.5%): {normal_data[4]:.4f}")
    print(f"Beta ({strategy_x} Lowest 2.5%): {tail_data[4]:.4f}")

    # Return the data for further analysis if needed
    return x_pos, y_pos, x_neg, y_neg, normal_data, tail_data
