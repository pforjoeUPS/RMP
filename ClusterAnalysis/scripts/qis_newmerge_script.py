# -*- coding: utf-8 -*-
"""
Created on Fri Jun 30 15:36:04 2023

@author: HQP6CBS
"""

import pandas as pd
import os
from functools import reduce
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
#from EquityHedging.datamanager import data_manager as dm
CWD = os.getcwd()


# Replace with the path to your original Excel file
existing_file = 'QIS Universe Time Series.xlsx'
existing_sheet_name = 'BofA'

# Replace  with the path to your new Excel file
new_file = 'update_QIS_uni.xlsx'
new_sheet_name = 'Sheet1'

# Read the original Excel file into a Pandas DataFrame
existing_df = pd.read_excel(CWD + "\\Cluster Analysis\\data\\" +existing_file, sheet_name=existing_sheet_name, skiprows=1, engine='openpyxl')

# Read the new data Excel file into a Pandas DataFrame
new_df = pd.read_excel(new_file, sheet_name=new_sheet_name)


# Get the ticker names from the second row of the new DataFrame (B2, C2, D2, etc.)
ticker_names_new = new_df.columns[1::2].tolist()

ticker_names_new = [f"{value} Index" for value in ticker_names_new]

# Rename the columns in the new DataFrame by adding the ' Index' suffix
new_df.rename(columns=dict(zip(new_df.columns[1::2], ticker_names_new)), inplace=True)

# Get the ticker names from the second row of the existing DataFrame (B2, C2, D2, etc.)
ticker_names_existing = existing_df.columns[2:].tolist()

# Count the number of tickers in the new data that match existing tickers
common_tickers_count = len(set(ticker_names_new).intersection(ticker_names_existing))

print("Number of tickers in the new data that match existing tickers:", common_tickers_count)

#Create a dictionary to store DataFrames for each ticker in the new data
new_strategy_data_frames = {}


# Combine the existing DataFrame with the new DataFrames
merged_df_combined = existing_df.copy()

ticker_data = {}

# Create a dictionary of DataFrames for each adjacent column pair in the new data
for i in range(0, new_df.shape[1], 2):
    date_column = new_df.iloc[1:, i].values.flatten()
    price_column = new_df.iloc[1:, i+1].values.flatten()

    # Create a DataFrame for the current adjacent column pair
    ticker_name = new_df.columns[i+1]
    ticker_df = pd.DataFrame({ticker_name: price_column}, index=date_column)

    # Convert the index to datetime type
    ticker_df.index = pd.to_datetime(ticker_df.index)
    ticker_df.index.name = 'Date'

    # Add the DataFrame to the dictionary
    ticker_data[ticker_name] = ticker_df

# Iterate through the DataFrames in the ticker_data dictionary
for ticker, ticker_df in ticker_data.items():
    if ticker in ticker_names_existing:
        # Replace existing ticker prices based on corresponding dates
        merged_df_combined[ticker] = merged_df_combined.apply(lambda row: ticker_df.loc[row['Date'], ticker] if row['Date'] in ticker_df.index else "", axis=1)
      
    else:
        # Left merge with existing DataFrame based on date
        merged_df_combined = pd.merge(merged_df_combined, ticker_df, left_on='Date', right_index=True, how='left')
        #merged_df_combined[ticker].fillna(value="NA", inplace=True)

# Fill the blanks with NaN

book = load_workbook(existing_file)

# Select the existing sheet
sheet = book[existing_sheet_name]

# Calculate the start row (2nd row in this case)
start_row = 2

# Convert the DataFrame to a list of rows
rows = dataframe_to_rows(merged_df_combined, index=False, header=True)

# Append the rows to the existing sheet starting from the 2nd row
for r_idx, row in enumerate(rows, start_row):
    for c_idx, value in enumerate(row, 1):
        sheet.cell(row=r_idx, column=c_idx, value=value)

# Save the updated Excel file
book.save(existing_file)
print("Data added to the existing sheet successfully!")
